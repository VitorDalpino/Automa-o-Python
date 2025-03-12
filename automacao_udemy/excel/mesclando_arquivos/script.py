from openpyxl import load_workbook, Workbook
import os

# Lista de arquivos Excel para serem consolidados
lista_arquivos = [
    r"G:\Meu Drive\MeusProjetos\projetos_py\automacao_udemy\excel\mesclando_arquivos\CustosAutom",
    r"G:\Meu Drive\MeusProjetos\projetos_py\automacao_udemy\excel\mesclando_arquivos\PopulacaoPOA",
    r"G:\Meu Drive\MeusProjetos\projetos_py\automacao_udemy\excel\mesclando_arquivos\SuperMercado"
]

wb = Workbook()
nome_arquivo_final = "resultado.xlsx"

# Removendo a aba inicial vazia
wb.remove(wb.active)

for nome_arquivo in lista_arquivos:
    caminho_arquivo = f"{nome_arquivo}.xlsx"

    try:
        arquivo = load_workbook(filename=caminho_arquivo)
        sheet = arquivo.active  # Pega a primeira aba ativa
        
        # Nome da aba no novo arquivo
        nome_aba = os.path.basename(nome_arquivo)[:31]  # Garante que o nome não ultrapasse 31 caracteres
        ws = wb.create_sheet(title=nome_aba)

        # Passando os dados de um arquivo para outro
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                ws.cell(row=i, column=j, value=sheet.cell(row=i, column=j).value)

    except Exception as e:
        print(f"Erro ao processar {caminho_arquivo}: {e}")

wb.save(nome_arquivo_final)
print(f"Consolidação concluída. Arquivo salvo como {nome_arquivo_final}.")
