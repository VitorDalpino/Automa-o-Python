from openpyxl import load_workbook

# Carregar o arquivo Excel
wb = load_workbook(filename=r"C:\Users\vitor\OneDrive\Documentos\MeusProjetos\projetos_py\automacao_udemy\excel\ler_arquivo\nomes.xlsx")
planilha = wb['Planilha1']

# Inicializar variável de soma
soma_idade = 0

# Loop para percorrer as linhas 2 a 4
for i in range(2, 5):
    nome = planilha[f'A{i}'].value
    idade = planilha[f'B{i}'].value
    
    # Verificar se o valor da idade não é None antes de converter
    if idade is not None:
        idade = int(idade)
        soma_idade += idade
        print(f"{nome} tem {idade} anos.")

# Atribuir o total da soma das idades na célula B5
planilha['B5'].value = soma_idade

# Mesclando e separando celulas
planilha['A7'] = "ALUNOS"
planilha.merge_cells("A7:B7")
# planilha.unmerge_cells("A7:B7")


# Salvar o arquivo Excel
wb.save(filename=r"C:\Users\vitor\OneDrive\Documentos\MeusProjetos\projetos_py\automacao_udemy\excel\ler_arquivo\nomes.xlsx")

print("Arquivo atualizado com sucesso!")
